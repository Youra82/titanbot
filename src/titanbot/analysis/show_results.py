# /root/titanbot/src/titanbot/analysis/show_results.py (Version für TitanBot SMC mit MaxDD - FEHLERFREI)
import os
import sys
import json
import pandas as pd
from datetime import date
import logging
import argparse 

# Logging etc. bleibt gleich
logging.getLogger('tensorflow').setLevel(logging.ERROR)
logging.getLogger('absl').setLevel(logging.ERROR)
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='keras')

PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', '..'))
sys.path.append(os.path.join(PROJECT_ROOT, 'src'))

from titanbot.analysis.backtester import load_data, run_smc_backtest
from titanbot.analysis.portfolio_simulator import run_portfolio_simulation
from titanbot.analysis.portfolio_optimizer import run_portfolio_optimizer
from titanbot.utils.telegram import send_document

GREEN  = '\033[0;32m'
YELLOW = '\033[1;33m'
NC     = '\033[0m'


def _get_telegram_cfg():
    try:
        with open(os.path.join(PROJECT_ROOT, 'secret.json'), 'r') as f:
            s = json.load(f)
        tg = s.get('telegram', {})
        return tg.get('bot_token', ''), tg.get('chat_id', '')
    except Exception:
        return '', ''


def _generate_trades_excel(final_sim, capital):
    """Erstellt titanbot_trades.xlsx mit allen Portfolio-Trades."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print(f"  {YELLOW}openpyxl nicht installiert — Excel übersprungen. (pip install openpyxl){NC}")
        return

    trade_history = final_sim.get('trade_history', [])
    if not trade_history:
        print(f"  {YELLOW}Keine Trades — Excel übersprungen.{NC}")
        return

    equity = capital
    rows   = []
    def _strip_tz(dt):
        """Entfernt Timezone-Info für Excel-Kompatibilität."""
        if hasattr(dt, 'tzinfo') and dt.tzinfo is not None:
            return dt.replace(tzinfo=None)
        return dt

    for i, t in enumerate(trade_history):
        pnl    = float(t['pnl'])
        symbol = t.get('symbol', '')
        tf     = t.get('timeframe', '')
        coin   = symbol.split('/')[0] if symbol else '—'
        dir_   = t.get('direction', '').upper()
        entry  = round(float(t.get('entry', 0)), 6)
        exit_p = round(float(t.get('exit',  0)), 6)
        ergebnis = 'TP erreicht' if pnl > 0 else 'SL erreicht'
        sl_pct   = t.get('sl_pct', 0)
        lev      = float(t.get('leverage', 1) or 1)
        margin   = float(t.get('margin_used', 0))
        if entry > 0:
            raw_move = (exit_p - entry) / entry * 100.0
            move_pct = raw_move if dir_ == 'LONG' else -raw_move
        else:
            move_pct = 0.0
        riskiert = round(margin * (sl_pct / 100.0) * lev, 4) if sl_pct else 0.0
        equity += pnl
        rows.append({
            'Nr':                 i + 1,
            'Datum':              _strip_tz(t.get('entry_time', '')),
            'Symbol':             coin,
            'Timeframe':          tf,
            'Richtung':           dir_,
            'Ergebnis':           ergebnis,
            'Reale Bewegung (%)': round(move_pct, 4),
            'Riskiert (USDT)':    riskiert,
            'Marge (USDT)':       round(margin, 4),
            'Hebel':              int(lev) if lev == int(lev) else round(lev, 1),
            'SL %':               f"{sl_pct:.3f}%" if sl_pct else '—',
            'Entry':              entry,
            'Exit':               exit_p,
            'PnL (USDT)':         round(pnl, 4),
            'Gesamtkapital':      round(equity, 4),
        })

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Trades'

    header_fill = PatternFill('solid', fgColor='1E3A5F')
    win_fill    = PatternFill('solid', fgColor='D6F4DC')
    loss_fill   = PatternFill('solid', fgColor='FAD7D7')
    alt_fill    = PatternFill('solid', fgColor='F2F2F2')
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),  bottom=Side(style='thin', color='CCCCCC'),
    )
    col_widths = {
        'Nr': 5, 'Datum': 18, 'Symbol': 10, 'Timeframe': 11, 'Richtung': 10,
        'Ergebnis': 14, 'Reale Bewegung (%)': 18, 'Riskiert (USDT)': 16, 'Marge (USDT)': 14,
        'Hebel': 8, 'SL %': 10, 'Entry': 14, 'Exit': 14, 'PnL (USDT)': 14, 'Gesamtkapital': 16,
    }

    headers = list(rows[0].keys())
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill      = header_fill
        cell.font      = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = thin_border
        ws.column_dimensions[get_column_letter(col)].width = col_widths.get(h, 14)
    ws.row_dimensions[1].height = 22

    for r_idx, row in enumerate(rows, 2):
        if row['Ergebnis'] == 'TP erreicht':
            fill = win_fill
        elif r_idx % 2 == 0:
            fill = loss_fill
        else:
            fill = alt_fill
        for col, key in enumerate(headers, 1):
            cell = ws.cell(row=r_idx, column=col, value=row[key])
            cell.fill      = fill
            cell.border    = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if key in ('Entry', 'Exit', 'PnL (USDT)', 'Gesamtkapital',
                       'Reale Bewegung (%)', 'Riskiert (USDT)', 'Marge (USDT)'):
                cell.number_format = '#,##0.0000'
        ws.row_dimensions[r_idx].height = 18

    total     = len(rows)
    wins      = sum(1 for r in rows if r['Ergebnis'] == 'TP erreicht')
    sr        = total + 3
    pnl_total = rows[-1]['Gesamtkapital'] - capital if rows else 0.0
    pnl_pct   = pnl_total / capital * 100 if capital else 0.0
    for label, value in [
        ('Trades gesamt', total),
        ('Win-Rate',      f"{wins / total * 100:.1f}%" if total else '—'),
        ('PnL',           f"{pnl_pct:+.1f}%"),
        ('Endkapital',    f"{rows[-1]['Gesamtkapital']:.2f} USDT" if rows else '—'),
    ]:
        ws.cell(row=sr, column=1, value=label).font = Font(bold=True)
        ws.cell(row=sr, column=2, value=value)
        sr += 1

    out_dir  = os.path.join(PROJECT_ROOT, 'artifacts', 'charts')
    os.makedirs(out_dir, exist_ok=True)
    out_file = os.path.join(out_dir, 'titanbot_trades.xlsx')
    wb.save(out_file)
    print(f"  {GREEN}Excel gespeichert: titanbot_trades.xlsx{NC}")

    bot_token, chat_id = _get_telegram_cfg()
    if bot_token and chat_id:
        caption = (f"TitanBot Trades — {total} Trades | "
                   f"WR: {wins / total * 100:.1f}% | PnL: {pnl_pct:+.1f}%" if total else "TitanBot Trades")
        send_document(bot_token, chat_id, out_file, caption=caption)
        print(f"  {GREEN}Via Telegram gesendet.{NC}")
    else:
        print(f"  {YELLOW}Telegram nicht konfiguriert — nur lokal gespeichert.{NC}")


def _generate_portfolio_chart(final_sim, portfolio_files, capital, start_date, end_date):
    """Erstellt titanbot_portfolio_equity.html (mbot-Style: Equity-Kurven + Entry/TP/SL)."""
    try:
        import plotly.graph_objects as go
        from plotly.subplots import make_subplots
    except ImportError:
        print(f"  {YELLOW}plotly nicht installiert - Chart uebersprungen.{NC}")
        return

    eq_df         = final_sim.get('equity_curve')
    trade_history = final_sim.get('trade_history', [])
    if eq_df is None or (hasattr(eq_df, 'empty') and eq_df.empty):
        print(f"  {YELLOW}Keine Equity-Daten - Chart uebersprungen.{NC}")
        return

    eq_idx = eq_df.index
    if hasattr(eq_idx, 'tz') and eq_idx.tz is not None:
        eq_idx = eq_idx.tz_convert('UTC').tz_localize(None)
    eq_times = [str(t)[:16] for t in eq_idx]
    eq_vals  = eq_df['equity'].tolist()

    def _ts_str(ts_raw):
        if ts_raw is None or str(ts_raw) == 'Backtest-Ende':
            return None
        return str(ts_raw)[:16].replace('T', ' ')

    entry_x, entry_y = [], []
    tp_x, tp_y       = [], []
    sl_x, sl_y       = [], []

    for t in trade_history:
        pnl        = float(t.get('pnl', 0))
        cap_after  = float(t.get('capital_after', capital + pnl))
        cap_before = cap_after - pnl
        entry_ts   = _ts_str(t.get('entry_time') or t.get('ts'))
        exit_ts    = _ts_str(t.get('exit_time'))
        if entry_ts:
            entry_x.append(entry_ts)
            entry_y.append(cap_before)
        marker_ts = exit_ts or entry_ts
        if not marker_ts:
            continue
        if pnl > 0:
            tp_x.append(marker_ts); tp_y.append(cap_after)
        else:
            sl_x.append(marker_ts); sl_y.append(cap_after)

    STRAT_COLORS = [
        '#f59e0b', '#10b981', '#8b5cf6', '#f97316',
        '#ec4899', '#14b8a6', '#a3e635', '#fb923c',
        '#e879f9', '#38bdf8',
    ]
    strat_trades = {}
    for t in trade_history:
        key = (t.get('strategy_key')
               or t.get('symbol', '').split('/')[0] + '/' + t.get('timeframe', ''))
        strat_trades.setdefault(key, []).append(t)

    pairs = []
    for fname in portfolio_files:
        name  = fname.replace('config_', '').replace('.json', '')
        parts = name.split('_')
        tf    = parts[-1] if parts else ''
        sym   = parts[0][:4].upper() if parts else ''
        pairs.append(f"{sym}/{tf}")
    pairs_str = ', '.join(pairs)

    n_strats = len(portfolio_files)
    pnl_pct  = final_sim.get('total_pnl_pct', 0)
    sign     = '+' if pnl_pct >= 0 else ''
    title = (
        f"TitanBot Portfolio - {n_strats} Strategie(n) ({pairs_str}) | "
        f"Zeitraum: {start_date} bis {end_date} | "
        f"Trades: {final_sim.get('trade_count', 0)} | WR: {final_sim.get('win_rate', 0):.1f}% | "
        f"PnL: {sign}{pnl_pct:.1f}% | "
        f"Endkapital: {final_sim.get('end_capital', capital):.2f} USDT | "
        f"MaxDD: {final_sim.get('max_drawdown_pct', 0):.1f}%"
    )

    fig = make_subplots(
        rows=2, cols=1,
        shared_xaxes=True,
        specs=[[{'secondary_y': False}], [{'secondary_y': False}]],
        vertical_spacing=0.03,
        row_heights=[0.85, 0.15],
    )

    fig.add_hline(
        y=capital,
        line=dict(color='rgba(150,150,150,0.4)', width=1, dash='dash'),
        annotation_text=f'Start {capital:.0f} USDT',
        annotation_position='top left',
        row=1, col=1,
    )

    for idx, (strat_key, trades) in enumerate(sorted(strat_trades.items())):
        eq  = capital
        xs  = [start_date + 'T00:00:00']
        ys  = [capital]
        for t in trades:
            eq += float(t.get('pnl', 0))
            x_t = (_ts_str(t.get('exit_time'))
                   or _ts_str(t.get('entry_time') or t.get('ts')))
            if x_t:
                xs.append(x_t)
                ys.append(round(eq, 4))
        color = STRAT_COLORS[idx % len(STRAT_COLORS)]
        sym   = strat_key.split('/')[0] if '/' in strat_key else strat_key[:6]
        tf    = strat_key.split('_')[-1] if '_' in strat_key else ''
        label = f"{sym}/{tf}" if tf else sym
        fig.add_trace(go.Scatter(
            x=xs, y=ys, mode='lines', name=label,
            line=dict(color=color, width=1.3, dash='dot'),
            opacity=0.65,
            hovertemplate=f"{label}: %{{y:.2f}} USDT<extra></extra>",
        ), row=1, col=1)

    fig.add_trace(go.Scatter(
        x=eq_times, y=eq_vals,
        mode='lines', name='Portfolio Equity',
        line=dict(color='#3b82f6', width=2.5),
        hovertemplate='Portfolio: %{y:.2f} USDT<extra></extra>',
    ), row=1, col=1)

    if entry_x:
        fig.add_trace(go.Scatter(
            x=entry_x, y=entry_y, mode='markers',
            marker=dict(symbol='triangle-up', size=10, color='#22c55e',
                        line=dict(width=1, color='#ffffff')),
            name='Entry',
            hovertemplate='Entry: %{x}<br>Kapital: %{y:.2f}<extra></extra>',
        ), row=1, col=1)

    if tp_x:
        fig.add_trace(go.Scatter(
            x=tp_x, y=tp_y, mode='markers',
            marker=dict(color='#22d3ee', symbol='circle', size=9,
                        line=dict(width=1, color='#0e7490')),
            name='TP',
            hovertemplate='TP: %{x}<br>Kapital: %{y:.2f}<extra></extra>',
        ), row=1, col=1)

    if sl_x:
        fig.add_trace(go.Scatter(
            x=sl_x, y=sl_y, mode='markers',
            marker=dict(color='#ef4444', symbol='x', size=9,
                        line=dict(width=2, color='#7f1d1d')),
            name='SL',
            hovertemplate='SL: %{x}<br>Kapital: %{y:.2f}<extra></extra>',
        ), row=1, col=1)

    if entry_x:
        fig.add_trace(go.Scatter(
            x=entry_x, y=[1]*len(entry_x), mode='markers',
            marker=dict(symbol='triangle-up', size=8, color='#22c55e'),
            showlegend=False,
            hovertemplate='Entry: %{x}<extra></extra>',
        ), row=2, col=1)
    if tp_x:
        fig.add_trace(go.Scatter(
            x=tp_x, y=[1]*len(tp_x), mode='markers',
            marker=dict(symbol='circle', size=7, color='#22d3ee'),
            showlegend=False,
            hovertemplate='TP: %{x}<extra></extra>',
        ), row=2, col=1)
    if sl_x:
        fig.add_trace(go.Scatter(
            x=sl_x, y=[1]*len(sl_x), mode='markers',
            marker=dict(symbol='x', size=8, color='#ef4444',
                        line=dict(width=2)),
            showlegend=False,
            hovertemplate='SL: %{x}<extra></extra>',
        ), row=2, col=1)

    fig.update_layout(
        title=dict(text=title, font=dict(size=12), x=0.5, xanchor='center'),
        height=720,
        hovermode='x unified',
        template='plotly_dark',
        dragmode='zoom',
        xaxis=dict(fixedrange=False),
        xaxis2=dict(fixedrange=False),
        legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='center', x=0.5),
        margin=dict(l=60, r=60, t=80, b=40),
        yaxis=dict(title='Equity (USDT)', fixedrange=False),
    )
    fig.update_yaxes(visible=False, row=2, col=1)

    out_dir  = os.path.join(PROJECT_ROOT, 'artifacts', 'charts')
    os.makedirs(out_dir, exist_ok=True)
    out_file = os.path.join(out_dir, 'titanbot_portfolio_equity.html')
    fig.write_html(out_file)
    print(f"  {GREEN}Chart gespeichert: titanbot_portfolio_equity.html{NC}")

    bot_token, chat_id = _get_telegram_cfg()
    if bot_token and chat_id:
        caption = (
            f"TitanBot Portfolio-Equity\n"
            f"{start_date} bis {end_date} | {n_strats} Strategie(n) | "
            f"PnL: {sign}{pnl_pct:.1f}% | Equity: {final_sim.get('end_capital', capital):.2f} USDT | "
            f"MaxDD: {final_sim.get('max_drawdown_pct', 0):.1f}%"
        )
        send_document(bot_token, chat_id, out_file, caption=caption)
        print(f"  {GREEN}Chart via Telegram gesendet.{NC}")
    else:
        print(f"  {YELLOW}Telegram nicht konfiguriert - Chart nur lokal gespeichert.{NC}")



# --- Einzel-Analyse ---
def run_single_analysis(start_date, end_date, start_capital, warmup_date=None, auto_write=False):
    print("--- TitanBot Ergebnis-Analyse (Einzel-Modus) ---")
    configs_dir = os.path.join(PROJECT_ROOT, 'src', 'titanbot', 'strategy', 'configs')
    all_results = []
    config_files = sorted([f for f in os.listdir(configs_dir) if f.startswith('config_') and f.endswith('.json')])
    if not config_files:
        print("\nKeine gültigen Konfigurationen zum Analysieren gefunden."); return
    data_start = warmup_date or start_date
    warmup_info = f" | SMC-Warmup ab {warmup_date}" if warmup_date else ""
    print(f"Backtest-Zeitraum: {start_date} bis {end_date} | Startkapital: {start_capital} USDT{warmup_info}")
    for filename in config_files:
        config_path = os.path.join(configs_dir, filename)
        if not os.path.exists(config_path): continue
        try:
            with open(config_path, 'r') as f: config = json.load(f)
            symbol = config['market']['symbol']
            timeframe = config['market']['timeframe']
            strategy_name = f"{symbol} ({timeframe})"
            print(f"\nAnalysiere: {filename}...")
            data = load_data(symbol, timeframe, data_start, end_date)
            if data.empty:
                print(f"--> WARNUNG: Keine Daten für {strategy_name}. Überspringe."); continue
            
            smc_params = config.get('strategy', {})
            risk_params = config.get('risk', {})
            
            # KORREKTUR: Füge Symbol, Timeframe und HTF zu smc_params hinzu, 
            # damit backtester.py die MTF-Logik anwenden kann.
            smc_params['symbol'] = symbol
            smc_params['timeframe'] = timeframe
            smc_params['htf'] = config['market'].get('htf')
            
            result = run_smc_backtest(data.copy(), smc_params, risk_params, start_capital, verbose=False, backtest_start_date=start_date if warmup_date else None)
            all_results.append({
                "file": filename,
                "symbol": symbol,
                "timeframe": timeframe,
                "Strategie": strategy_name,
                "Trades": result.get('trades_count', 0),
                "Win Rate %": result.get('win_rate', 0),
                "PnL %": result.get('total_pnl_pct', -100),
                "Max DD %": result.get('max_drawdown_pct', 1.0) * 100,
                "Endkapital": result.get('end_capital', start_capital)
            })
        except Exception as e:
            print(f"--> FEHLER bei {filename}: {e}")
            continue
    if not all_results:
        print("\nKeine gültigen Ergebnisse zum Anzeigen gefunden."); return
    results_df = pd.DataFrame(all_results)
    results_df = results_df.sort_values(by="PnL %", ascending=False)
    pd.set_option('display.width', 1000); pd.set_option('display.max_columns', None)
    print("\n\n=========================================================================================");
    print(f"                        Zusammenfassung aller Einzelstrategien");
    print("=========================================================================================")
    pd.set_option('display.float_format', '{:.2f}'.format)
    display_cols = ["Strategie", "Trades", "Win Rate %", "PnL %", "Max DD %", "Endkapital"]
    print(results_df[display_cols].to_string(index=False))
    print("=========================================================================================")

    # --- Auto-Write: Top-Strategien mit positivem PnL in settings.json eintragen ---
    if auto_write:
        settings_path = os.path.join(PROJECT_ROOT, "settings.json")
        try:
            with open(settings_path, "r") as f:
                settings = json.load(f)
            max_pos = int(settings.get("live_trading_settings", {}).get("max_open_positions", 5))
            positive = [r for r in all_results if r["PnL %"] > 0]
            top = sorted(positive, key=lambda x: x["PnL %"], reverse=True)[:max_pos]
            if top:
                new_active = [{"symbol": r["symbol"], "timeframe": r["timeframe"], "active": True} for r in top]
                settings["live_trading_settings"]["active_strategies"] = new_active
                with open(settings_path, "w") as f:
                    json.dump(settings, f, indent=4)
                print(f"\n{GREEN}\u2705 {len(new_active)} Strategien in settings.json eingetragen:{NC}")
                for s in new_active:
                    print(f"   - {s['symbol']} ({s['timeframe']})")
            else:
                print(f"\n{YELLOW}\u26a0  Keine Strategie mit positivem PnL - settings.json unveraendert.{NC}")
        except Exception as e:
            print(f"\nFEHLER beim Schreiben in settings.json: {e}")


# --- Geteilter Modus (Manuell / Auto) ---
def run_shared_mode(is_auto: bool, start_date, end_date, start_capital, target_max_dd: float, warmup_date=None):
    mode_name = "Automatische Portfolio-Optimierung" if is_auto else "Manuelle Portfolio-Simulation"
    data_start = warmup_date or start_date
    print(f"--- TitanBot {mode_name} ---")
    if is_auto:
        print(f"Ziel: Maximaler Profit bei maximal {target_max_dd:.2f}% Drawdown.")

    configs_dir = os.path.join(PROJECT_ROOT, 'src', 'titanbot', 'strategy', 'configs')
    available_strategies = []
    if os.path.isdir(configs_dir):
        for filename in sorted(os.listdir(configs_dir)):
            if filename.startswith('config_') and filename.endswith('.json'):
                available_strategies.append(filename)
    if not available_strategies:
        print("Keine optimierten Strategien (Configs) gefunden."); return

    selected_files = []
    if not is_auto:
        print("\nVerfügbare Strategien:")
        for i, name in enumerate(available_strategies): print(f"  {i+1}) {name}")
        selection = input("\nWelche Strategien sollen simuliert werden? (Zahlen mit Komma, z.B. 1,3,4 oder 'alle'): ")
        try:
            if selection.lower() == 'alle': selected_files = available_strategies
            else: selected_files = [available_strategies[int(i.strip()) - 1] for i in selection.split(',')]
        except (ValueError, IndexError): print("Ungültige Auswahl. Breche ab."); return
    else:
        selected_files = available_strategies

    strategies_data = {}
    print("\nLade Daten für gewählte Strategien...")
    for filename in selected_files:
        try:
            with open(os.path.join(configs_dir, filename), 'r') as f: config = json.load(f)
            symbol = config['market']['symbol']
            timeframe = config['market']['timeframe']
            # Lese HTF aus der Konfiguration
            htf = config['market'].get('htf')
            
            data = load_data(symbol, timeframe, data_start, end_date)
            if not data.empty:
                strategies_data[filename] = {
                    'symbol': symbol, 'timeframe': timeframe, 'data': data,
                    'smc_params': config.get('strategy', {}),
                    'risk_params': config.get('risk', {}),
                    'htf': htf 
                }
            else:
                print(f"WARNUNG: Konnte Daten für {filename} nicht laden. Wird ignoriert.")
        except Exception as e:
            print(f"FEHLER beim Laden der Config/Daten für {filename}: {e}")

    if not strategies_data:
        print("Konnte für keine der gewählten Strategien Daten laden. Breche ab."); return

    final_sim = None
    portfolio_files_used = selected_files

    try:
        if is_auto:
            results = run_portfolio_optimizer(start_capital, strategies_data, start_date, end_date, target_max_dd)

            if results and "final_result" in results and results["final_result"] is not None:
                final_report = results["final_result"]
                print("\n======================================================="); print("     Ergebnis der automatischen Portfolio-Optimierung"); print("=======================================================")
                print(f"Zeitraum: {start_date} bis {end_date}\nStartkapital: {start_capital:.2f} USDT")
                print(f"Bedingung: Max Drawdown <= {target_max_dd:.2f}%")

                if results.get("optimal_portfolio"):
                    portfolio_files_used = results["optimal_portfolio"]
                    print("\nOptimales Portfolio gefunden (" + str(len(portfolio_files_used)) + " Strategien):")
                    for strat_filename in portfolio_files_used: print(f"  - {strat_filename}")
                else:
                    print("\nBeste Einzelstrategie gefunden:")
                    strat_key = final_report.get("strategy_key", "Unbekannt")
                    print(f"  - {strat_key}")

                print("\n--- Simulierte Performance dieses Portfolios/dieser Strategie ---")
                print(f"Endkapital:         {final_report['end_capital']:.2f} USDT"); print(f"Gesamt PnL:         {final_report['end_capital'] - start_capital:+.2f} USDT ({final_report['total_pnl_pct']:.2f}%)")
                print(f"Portfolio Max DD:   {final_report['max_drawdown_pct']:.2f}%")
                liq_date = final_report.get("liquidation_date")
                print(f"Liquidiert:         {'JA, am ' + liq_date.strftime('%Y-%m-%d') if liq_date else 'NEIN'}")
                final_sim = final_report
            else:
                print(f"\nKein Portfolio gefunden, das die Bedingung Max Drawdown <= {target_max_dd:.2f}% erfuellt.")

        # --- Manuelle Simulation ---
        else:
            sim_data = {v["symbol"] + "_" + v["timeframe"]: v for k, v in strategies_data.items()}
            results = run_portfolio_simulation(start_capital, sim_data, start_date, end_date)
            if results:
                print("\n======================================================="); print("           Portfolio-Simulations-Ergebnis"); print("=======================================================")
                print(f"Zeitraum: {start_date} bis {end_date}\nStartkapital: {results['start_capital']:.2f} USDT")
                print("\n--- Gesamt-Performance ---")
                print(f"Endkapital:         {results['end_capital']:.2f} USDT"); print(f"Gesamt PnL:         {results['end_capital'] - results['start_capital']:+.2f} USDT ({results['total_pnl_pct']:.2f}%)")
                print(f"Anzahl Trades:      {results['trade_count']}"); print(f"Win-Rate:           {results['win_rate']:.2f}%")
                print(f"Portfolio Max DD:   {results['max_drawdown_pct']:.2f}% am {results['max_drawdown_date'].strftime('%Y-%m-%d') if results['max_drawdown_date'] else 'N/A'}")
                liq_date = results.get("liquidation_date")
                print(f"Liquidiert:         {'JA, am ' + liq_date.strftime('%Y-%m-%d') if liq_date else 'NEIN'}")
                final_sim = results

    except Exception as e:
        print(f"\nFEHLER waehrend der Portfolio-Analyse: {e}")
        import traceback
        traceback.print_exc()

    # --- Export: Excel + HTML Chart ---
    if final_sim is not None:
        print("\n--- Export ---")
        _generate_trades_excel(final_sim, start_capital)
        _generate_portfolio_chart(final_sim, portfolio_files_used, start_capital, start_date, end_date)
    else:
        print("\nPortfolio-Analyse fehlgeschlagen oder kein gueltiges Portfolio gefunden.")



if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument('--mode', default='1', type=str, help="Analyse-Modus (1=Einzel, 2=Manuell, 3=Auto, 4=Interaktiv)")
    parser.add_argument('--target_max_drawdown', default=30.0, type=float, help="Ziel Max Drawdown % (nur für Modus 3)")
    parser.add_argument('--start_date', default='2023-01-01', type=str, help="Startdatum JJJJ-MM-TT")
    parser.add_argument('--end_date', default=None, type=str, help="Enddatum JJJJ-MM-TT (Standard: Heute)")
    parser.add_argument('--start_capital', default=1000, type=int, help="Startkapital in USDT")
    parser.add_argument('--warmup_date', default=None, type=str, help="Datum ab dem Daten fuer SMC-Warmup geladen werden")
    parser.add_argument('--auto_write', action='store_true', help="Beste Strategien automatisch in settings.json schreiben")
    args = parser.parse_args()

    start_date = args.start_date
    end_date = args.end_date or date.today().strftime("%Y-%m-%d")
    start_capital = args.start_capital


    # --- OOS-Info: Zeige Training/OOS-Aufteilung je Timeframe ---
    try:
        import json as _json
        from datetime import date as _date, timedelta as _td
        _settings_path = os.path.join(PROJECT_ROOT, 'settings.json')
        with open(_settings_path) as _sf:
            _settings = _json.load(_sf)
        _oos_ref = _settings.get('optimization_settings', {}).get('oos_reference_date')
        if _oos_ref:
            _configs_dir = os.path.join(PROJECT_ROOT, 'src', 'titanbot', 'strategy', 'configs')
            _cfg_tfs = set()
            for _fn in os.listdir(_configs_dir):
                if _fn.startswith('config_') and _fn.endswith('.json'):
                    try:
                        with open(os.path.join(_configs_dir, _fn)) as _cf:
                            _cfg_tfs.add(_json.load(_cf)['market']['timeframe'])
                    except Exception:
                        pass
            _LOOKBACK_MAP = {'5m': 60, '15m': 60, '30m': 365, '1h': 365,
                             '2h': 730, '4h': 730, '6h': 730, '1d': 1095}
            _tfs_to_check = {_tf: _lb for _tf, _lb in _LOOKBACK_MAP.items()
                             if not _cfg_tfs or _tf in _cfg_tfs}
            _ref_dt         = _date.fromisoformat(str(_oos_ref))
            _analysis_start = _date.fromisoformat(start_date)
            _analysis_end   = _date.fromisoformat(end_date)
            _total_days     = (_analysis_end - _analysis_start).days + 1
            print()
            print(f"  OOS-Referenz: {_oos_ref}  |  Analysezeitraum: {start_date} → {end_date} ({_total_days} Tage)")
            print(f"  {'TF':>4s}  {'OOS ab':>12s}  {'Training':>10s}  {'OOS':>8s}  Status")
            print(f"  {'─'*4}  {'─'*12}  {'─'*10}  {'─'*8}  {'─'*30}")
            for _tf, _lb in sorted(_tfs_to_check.items()):
                _oos_days_tf = _lb * 30 // 100
                _oos_start   = _ref_dt - _td(days=_oos_days_tf)
                _train_end   = _oos_start - _td(days=1)
                _t_days = (min(_analysis_end, _train_end) - _analysis_start).days + 1 if _analysis_start <= _train_end else 0
                _o_days = (_analysis_end - max(_analysis_start, _oos_start)).days + 1 if _analysis_end >= _oos_start else 0
                if _t_days <= 0:
                    _status = "✅ vollständig OOS"
                elif _o_days <= 0:
                    _status = "⚠️  vollständig Training"
                else:
                    _pct_train = _t_days * 100 // _total_days
                    _status = f"ℹ️  {_t_days}d Training / {_o_days}d OOS ({_pct_train}% im Training)"
                print(f"  {_tf:>4s}  {str(_oos_start):>12s}  {max(0,_t_days):>9d}d  {max(0,_o_days):>7d}d  {_status}")
            print()
    except Exception:
        pass  # OOS-Pruefung optional
    if args.mode == '4':
        # Modus 4: Interaktive Charts (SMC)
        print("\n--- Starte interaktive Chart-Generierung (SMC) ---")
        from titanbot.analysis.interactive_status import main as interactive_main
        interactive_main()
    else:
        print("--------------------------------------------------")

        if args.mode == '2':
            # KORREKTUR: Explizite Benennung der Argumente für den run_shared_mode Aufruf (behebt TypeError)
            run_shared_mode(
                is_auto=False, 
                start_date=start_date, 
                end_date=end_date, 
                start_capital=start_capital, 
                target_max_dd=999.0
            )
        elif args.mode == '3':
            # KORREKTUR: Explizite Benennung der Argumente für den run_shared_mode Aufruf (behebt TypeError)
            run_shared_mode(
                is_auto=True, 
                start_date=start_date, 
                end_date=end_date, 
                start_capital=start_capital, 
                target_max_dd=args.target_max_drawdown
            )
        else: # Modus 1 (default)
            run_single_analysis(start_date=start_date, end_date=end_date, start_capital=start_capital,
                                warmup_date=args.warmup_date, auto_write=args.auto_write)
