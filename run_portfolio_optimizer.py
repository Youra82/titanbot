#!/usr/bin/env python3
"""
run_portfolio_optimizer.py  (titanbot)

Lädt alle Configs, führt Portfolio-Simulation (gemeinsamer Kapital-Pool,
kombinierte Equity-Kurve, echter MaxDD) durch und wählt das beste Portfolio
per Greedy-Algorithmus. Schreibt active_strategies in settings.json.

Aufruf:
  python3 run_portfolio_optimizer.py              # interaktiv
  python3 run_portfolio_optimizer.py --auto-write # automatisch (Scheduler)
"""
import contextlib
import io
import os
import sys
import json
import argparse
from datetime import date, timedelta
from tqdm import tqdm

PROJECT_ROOT  = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(PROJECT_ROOT, 'src'))

CONFIGS_DIR   = os.path.join(PROJECT_ROOT, 'src', 'titanbot', 'strategy', 'configs')
SETTINGS_PATH = os.path.join(PROJECT_ROOT, 'settings.json')

B  = '\033[1;37m'
G  = '\033[0;32m'
Y  = '\033[1;33m'
R  = '\033[0;31m'
NC = '\033[0m'

DEFAULT_LOOKBACK_DAYS = 1095  # ~3 Jahre als Standard


def _scan_configs() -> list:
    if not os.path.isdir(CONFIGS_DIR):
        return []
    return sorted([
        os.path.join(CONFIGS_DIR, f)
        for f in os.listdir(CONFIGS_DIR)
        if f.endswith('.json')
    ])


def _build_strategies_data(config_files: list, start_date: str, end_date: str) -> dict:
    from titanbot.analysis.backtester import load_data
    strategies_data = {}
    for path in tqdm(config_files, desc='Lade Configs & Daten'):
        fname = os.path.basename(path)
        try:
            with open(path) as f:
                config = json.load(f)
            market    = config.get('market', {})
            symbol    = market.get('symbol', '')
            timeframe = market.get('timeframe', '')
            htf       = market.get('htf')
            if not symbol or not timeframe:
                continue
            data = load_data(symbol, timeframe, start_date, end_date)
            if data is None or data.empty or len(data) < 50:
                print(f"  {Y}Uebersprungen (keine Daten): {fname}{NC}")
                continue
            strategies_data[fname] = {
                'symbol':     symbol,
                'timeframe':  timeframe,
                'data':       data,
                'smc_params': config.get('strategy', {}),
                'risk_params': config.get('risk', {}),
                'htf':        htf,
            }
        except Exception as e:
            print(f"  {Y}Fehler bei {fname}: {e}{NC}")
    return strategies_data


def _simulate_current_portfolio(settings: dict, strategies_data: dict,
                                 start_capital: float,
                                 start_date: str, end_date: str) -> dict | None:
    """Simuliert das aktuell aktive Portfolio auf dem gleichen Zeitraum."""
    from titanbot.analysis.portfolio_simulator import run_portfolio_simulation
    current = [
        s for s in settings.get('live_trading_settings', {}).get('active_strategies', [])
        if s.get('active')
    ]
    if not current:
        return None
    sim_data = {}
    for s in current:
        sym, tf = s.get('symbol', ''), s.get('timeframe', '')
        for fname, sd in strategies_data.items():
            if sd['symbol'] == sym and sd['timeframe'] == tf:
                sim_data[f"{sym}_{tf}"] = sd
                break
    if not sim_data:
        return None
    with contextlib.redirect_stdout(io.StringIO()), contextlib.redirect_stderr(io.StringIO()):
        return run_portfolio_simulation(start_capital, sim_data, start_date, end_date)


def _write_to_settings(portfolio_files: list, strategies_data: dict) -> None:
    with open(SETTINGS_PATH) as f:
        settings = json.load(f)
    existing     = settings.get('live_trading_settings', {}).get('active_strategies', [])
    existing_map = {(s.get('symbol'), s.get('timeframe')): s for s in existing}
    new_strategies = []
    for fname in portfolio_files:
        sd        = strategies_data.get(fname, {})
        symbol    = sd.get('symbol', '')
        timeframe = sd.get('timeframe', '')
        if not symbol or not timeframe:
            continue
        base  = existing_map.get((symbol, timeframe), {})
        entry = {**base, 'symbol': symbol, 'timeframe': timeframe, 'active': True}
        new_strategies.append(entry)
    lt = settings.setdefault('live_trading_settings', {})
    lt['active_strategies']          = new_strategies
    lt['use_auto_optimizer_results'] = True
    with open(SETTINGS_PATH, 'w') as f:
        json.dump(settings, f, indent=4)


BOT_NAME = 'titanbot'


def _get_telegram_creds():
    try:
        with open(os.path.join(PROJECT_ROOT, 'secret.json')) as f:
            s = json.load(f)
        tg = s.get('telegram', {})
        t, c = tg.get('bot_token', ''), tg.get('chat_id', '')
        return (t, c) if t and c else (None, None)
    except Exception:
        return None, None


def _send_telegram(msg):
    token, chat = _get_telegram_creds()
    if not token:
        return
    try:
        import requests
        requests.post(f'https://api.telegram.org/bot{token}/sendMessage',
                      data={'chat_id': chat, 'text': msg}, timeout=10)
    except Exception:
        pass


def _send_telegram_doc(fpath, caption=''):
    token, chat = _get_telegram_creds()
    if not token:
        return
    try:
        import requests
        with open(fpath, 'rb') as fh:
            requests.post(f'https://api.telegram.org/bot{token}/sendDocument',
                          data={'chat_id': chat, 'caption': caption},
                          files={'document': fh}, timeout=30)
    except Exception:
        pass


def generate_trades_excel(final, strategies_data, capital, start_date, end_date):
    """Erstellt Excel-Tabelle mit allen Portfolio-Trades."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print(f'  {Y}openpyxl nicht installiert — Excel uebersprungen.{NC}')
        return None

    trades = final.get('trade_history', [])
    if not trades:
        return None

    equity = capital
    rows = []
    for i, t in enumerate(trades, 1):
        pnl = t.get('pnl', 0.0)
        equity += pnl
        rows.append({
            'Nr':            i,
            'Datum':         str(t.get('entry_time', t.get('ts', '')))[:16].replace('T', ' '),
            'Symbol':        t.get('symbol', '?'),
            'Timeframe':     t.get('timeframe', '?'),
            'Richtung':      str(t.get('direction', '?')).upper(),
            'Ergebnis':      'TP erreicht' if pnl >= 0 else 'SL erreicht',
            'PnL (USDT)':    round(pnl, 4),
            'Gesamtkapital': round(equity, 4),
        })

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Trades'
    hdr  = PatternFill('solid', fgColor='1E3A5F')
    win  = PatternFill('solid', fgColor='D6F4DC')
    loss = PatternFill('solid', fgColor='FAD7D7')
    alt  = PatternFill('solid', fgColor='F2F2F2')
    brd  = Border(left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
                  top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))
    cw   = {'Nr': 6, 'Datum': 18, 'Symbol': 22, 'Timeframe': 12, 'Richtung': 10,
             'Ergebnis': 14, 'PnL (USDT)': 14, 'Gesamtkapital': 16}
    hdrs = list(rows[0].keys())
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = hdr
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = brd
        ws.column_dimensions[get_column_letter(c)].width = cw.get(h, 14)
    ws.row_dimensions[1].height = 22
    for ri, row in enumerate(rows, 2):
        f = win if row['Ergebnis'] == 'TP erreicht' else (loss if ri % 2 == 0 else alt)
        for c, key in enumerate(hdrs, 1):
            cell = ws.cell(row=ri, column=c, value=row[key])
            cell.fill = f
            cell.border = brd
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if key in ('PnL (USDT)', 'Gesamtkapital'):
                cell.number_format = '#,##0.0000'
        ws.row_dimensions[ri].height = 18
    pnl = final.get('total_pnl_pct', 0)
    dd  = final.get('max_drawdown_pct', 0)
    wr  = final.get('win_rate', 0)
    eq  = final.get('end_capital', equity)
    n   = final.get('trade_count', len(trades))
    sr  = len(rows) + 3
    for label, val in [('Zeitraum', f'{start_date} -> {end_date}'), ('Trades', n),
                        ('Win-Rate', f'{wr:.1f}%'), ('PnL', f'{pnl:+.1f}%'),
                        ('Endkapital', f'{eq:.2f} USDT'), ('Max Drawdown', f'{dd:.1f}%')]:
        ws.cell(row=sr, column=1, value=label).font = Font(bold=True)
        ws.cell(row=sr, column=2, value=val)
        sr += 1
    outfile = f'/tmp/{BOT_NAME}_trades.xlsx'
    wb.save(outfile)
    print(f'  {G}✓ Excel erstellt: {outfile}{NC}')
    return outfile


def generate_equity_html(final, capital, start_date, end_date, labels):
    """Erstellt interaktiven Portfolio-Equity-Chart."""
    try:
        import plotly.graph_objects as go
    except ImportError:
        print(f'  {Y}plotly nicht installiert — Chart uebersprungen.{NC}')
        return None

    eq_df = final.get('equity_curve')
    if eq_df is None or (hasattr(eq_df, 'empty') and eq_df.empty):
        return None

    times = [str(t) for t in eq_df['timestamp']]
    vals  = [float(v) for v in eq_df['equity']]
    pnl   = final.get('total_pnl_pct', 0)
    dd    = final.get('max_drawdown_pct', 0)
    wr    = final.get('win_rate', 0)
    n     = final.get('trade_count', 0)
    eq    = final.get('end_capital', vals[-1] if vals else capital)
    sign  = '+' if pnl >= 0 else ''
    title = (f"{BOT_NAME} Portfolio — {', '.join(labels)} | "
             f"PnL: {sign}{pnl:.1f}% | Equity: {eq:.2f} USDT | "
             f"MaxDD: {dd:.1f}% | WR: {wr:.1f}% | {n} Trades")

    fig = go.Figure()
    fig.add_trace(go.Scatter(x=times, y=vals, mode='lines', name='Portfolio Equity',
                             line=dict(color='#2563eb', width=2)))
    fig.add_hline(y=capital, line=dict(color='rgba(100,100,100,0.4)', width=1, dash='dash'),
                  annotation_text=f'Start {capital:.0f} USDT', annotation_position='top left')
    fig.update_layout(title=dict(text=title, font=dict(size=12), x=0.5),
                      height=600, template='plotly_white', hovermode='x unified',
                      xaxis=dict(rangeslider=dict(visible=True), fixedrange=False),
                      yaxis=dict(title='Equity (USDT)', fixedrange=False))
    outfile = f'/tmp/{BOT_NAME}_portfolio_equity.html'
    fig.write_html(outfile)
    print(f'  {G}✓ Chart erstellt: {outfile}{NC}')
    return outfile


def main() -> int:
    parser = argparse.ArgumentParser(description='titanbot Portfolio-Optimizer')
    parser.add_argument('--capital',    type=float, default=None)
    parser.add_argument('--max-dd',     type=float, default=30.0)
    parser.add_argument('--start-date', type=str,   default=None)
    parser.add_argument('--end-date',   type=str,   default=None)
    parser.add_argument('--auto-write', action='store_true')
    args = parser.parse_args()

    with open(SETTINGS_PATH) as f:
        settings = json.load(f)
    opt           = settings.get('optimization_settings', {})
    capital       = args.capital or float(opt.get('start_capital', 100))
    max_dd        = args.max_dd
    end_date      = args.end_date   or date.today().strftime('%Y-%m-%d')
    start_date    = args.start_date or (
        date.today() - timedelta(days=DEFAULT_LOOKBACK_DAYS)
    ).strftime('%Y-%m-%d')
    max_positions = int(settings.get('live_trading_settings', {}).get('max_open_positions', 5))

    print(f"\n{'─'*72}")
    print(f"{B}  titanbot — Automatische Portfolio-Optimierung{NC}")
    print(f"  Greedy-Selektion mit echter Portfolio-Simulation (MaxDD ≤ {max_dd:.0f}%)")
    print(f"  Kapital: {capital:.0f} USDT | Positionen: max {max_positions} | "
          f"Zeitraum: {start_date} → {end_date}")
    print(f"{'─'*72}\n")

    config_files = _scan_configs()
    if not config_files:
        print(f"{R}  Keine Configs in {CONFIGS_DIR}{NC}")
        print(f"  → Zuerst run_pipeline.sh ausfuehren!\n")
        return 1

    print(f"  {len(config_files)} Config(s) gefunden.\n")
    strategies_data = _build_strategies_data(config_files, start_date, end_date)
    if not strategies_data:
        print(f"{R}  Keine Daten geladen.{NC}")
        return 1

    from titanbot.analysis.portfolio_optimizer import run_portfolio_optimizer
    result = run_portfolio_optimizer(capital, strategies_data, start_date, end_date, max_dd)

    if not result or not result.get('optimal_portfolio'):
        print(f"{R}  Kein Portfolio erfuellt die Bedingungen (MaxDD ≤ {max_dd:.0f}%).{NC}\n")
        return 0

    portfolio_files = result['optimal_portfolio'][:max_positions]
    final           = result.get('final_result') or {}

    print(f"\n{'='*72}")
    print(f"{B}  Optimales Portfolio — {len(portfolio_files)} Strategie(n){NC}\n")
    for fname in portfolio_files:
        sd = strategies_data.get(fname, {})
        print(f"  {G}✓{NC} {sd.get('symbol', fname):<26} / {sd.get('timeframe', ''):<6}")
    if final:
        pnl = final.get('total_pnl_pct', 0)
        print(f"\n  Endkapital: {final.get('end_capital', 0):.2f} USDT  "
              f"| PnL: {pnl:+.1f}%  "
              f"| MaxDD: {final.get('max_drawdown_pct', 0):.2f}%")
    print(f"{'='*72}\n")

    current_set = {
        (s.get('symbol'), s.get('timeframe'))
        for s in settings.get('live_trading_settings', {}).get('active_strategies', [])
        if s.get('active')
    }
    new_set = {
        (strategies_data.get(f, {}).get('symbol'), strategies_data.get(f, {}).get('timeframe'))
        for f in portfolio_files
    }

    cur_result  = _simulate_current_portfolio(settings, strategies_data, capital, start_date, end_date)
    cur_cap     = cur_result.get('end_capital', 0) if cur_result else 0
    new_cap     = final.get('end_capital', 0)
    if cur_result:
        print(f"  Aktuelles Portfolio: {cur_cap:.2f} USDT  "
              f"| PnL: {cur_result.get('total_pnl_pct', 0):+.1f}%  "
              f"| MaxDD: {cur_result.get('max_drawdown_pct', 0):.2f}%")
        print(f"  Neues Portfolio:     {new_cap:.2f} USDT  "
              f"| PnL: {final.get('total_pnl_pct', 0):+.1f}%  "
              f"| MaxDD: {final.get('max_drawdown_pct', 0):.2f}%\n")

    if args.auto_write:
        if cur_result and new_cap <= cur_cap:
            print(f"{Y}  Neues Portfolio ({new_cap:.2f} USDT) nicht besser als aktuelles "
                  f"({cur_cap:.2f} USDT) — keine Aenderung.{NC}\n")
        else:
            _write_to_settings(portfolio_files, strategies_data)
            print(f"{G}✓ settings.json aktualisiert — {len(portfolio_files)} Strategie(n).{NC}\n")
    else:
        if current_set == new_set:
            print(f"{Y}  Portfolio unveraendert — keine Aenderung noetig.{NC}\n")
        else:
            try:
                ans = input("  Optimales Portfolio in settings.json eintragen? (j/n): ").strip().lower()
            except (EOFError, KeyboardInterrupt):
                ans = 'n'
            if ans in ('j', 'ja', 'y', 'yes'):
                _write_to_settings(portfolio_files, strategies_data)
                print(f"{G}✓ settings.json aktualisiert.{NC}\n")
            else:
                print(f"{Y}  settings.json NICHT geaendert.{NC}\n")

    # ── Reports & Telegram ──────────────────────────────────────────────────
    if args.auto_write:
        labels = [
            f"{strategies_data.get(f, {}).get('symbol', '?')}/{strategies_data.get(f, {}).get('timeframe', '?')}"
            for f in portfolio_files
        ]
        pnl = final.get('total_pnl_pct', 0)
        dd  = final.get('max_drawdown_pct', 0)
        n   = final.get('trade_count', 0)
        wr  = final.get('win_rate', 0)
        eq  = final.get('end_capital', 0)
        summary = (f"{BOT_NAME} Auto-Optimizer\n"
                   f"{len(portfolio_files)} Strategien | {n} Trades | WR: {wr:.1f}%\n"
                   f"PnL: {pnl:+.1f}% | MaxDD: {dd:.1f}% | Equity: {eq:.2f} USDT\n"
                   f"Zeitraum: {start_date} -> {end_date}")
        _send_telegram(summary)
        xlsx = generate_trades_excel(final, strategies_data, capital, start_date, end_date)
        if xlsx:
            _send_telegram_doc(xlsx, caption=f'{BOT_NAME} Trades | {n} Trades | WR: {wr:.1f}% | Equity: {eq:.2f} USDT')
        html = generate_equity_html(final, capital, start_date, end_date, labels)
        if html:
            _send_telegram_doc(html, caption=f'{BOT_NAME} Portfolio-Equity | PnL: {pnl:+.1f}% | MaxDD: {dd:.1f}%')

    return 0


if __name__ == '__main__':
    sys.exit(main())
